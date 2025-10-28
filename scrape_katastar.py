#!/usr/bin/env python3
import sys
import argparse
import re
from typing import List, Tuple, Optional

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# -----------------------------
# Selenium helpers
# -----------------------------
def get_region_suggestions(driver, wait, region_text):
    """
    Types region_text into the 'Внеси катастарска општина' input and returns all visible suggestions <li role='option'>.
    If none show up within timeout, returns [].
    """
    region_input = wait.until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, "input[placeholder='Внеси катастарска општина']")))
    region_input.clear()
    region_input.send_keys(region_text)

    try:
        wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "li[role='option']")))
    except TimeoutException:
        return []

    return driver.find_elements(By.CSS_SELECTOR, "li[role='option']")


def select_parcel(driver, wait, parcel_text):
    """
    Enters parcel_text into 'Внеси парцела', clicks a suggestion if present.
    Returns True if a suggestion was clicked; otherwise False.
    """
    parcel_input = wait.until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, "input[placeholder='Внеси парцела']")))
    parcel_input.clear()
    parcel_input.send_keys(parcel_text)

    try:
        wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "li[role='option']")))
    except TimeoutException:
        return False

    suggestions = driver.find_elements(By.CSS_SELECTOR, "li[role='option']")
    chosen = None
    for s in suggestions:
        if parcel_text in s.text:
            chosen = s
            break
    if chosen is None:
        chosen = suggestions[0]

    driver.execute_script("arguments[0].click();", chosen)
    return True


def extract_parcel_and_holders(driver, wait):
    """
    On the parcel results page, extract the parcels table and for each parcel open the right-holders details.
    Returns a list of {parcel fields..., 'Носители на право': [holders...]} dicts.
    """
    wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tr.parcels-table-body-row")))
    rows_count = len(driver.find_elements(By.CSS_SELECTOR, "tr.parcels-table-body-row"))
    result = []

    for i in range(rows_count):
        parcel_rows = driver.find_elements(By.CSS_SELECTOR, "tr.parcels-table-body-row")
        row = parcel_rows[i]
        cells = row.find_elements(By.TAG_NAME, "td")

        data = {
            "Имотен лист": cells[1].text,
            "Број/дел": cells[2].text,
            "Култура": cells[3].text,
            "Површина m2": cells[4].text,
            "Место": cells[5].text,
            "Право": cells[6].text,
        }

        original_handle = driver.current_window_handle
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cells[1])
        driver.execute_script("arguments[0].click();", cells[1])

        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#right-holders-table-paper")))

        if len(driver.window_handles) > 1:
            for handle in driver.window_handles:
                if handle != original_handle:
                    driver.switch_to.window(handle)
                    break

        holder_rows = wait.until(EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "#right-holders-table-paper tbody tr")))
        holders = []
        for hrow in holder_rows:
            hc = hrow.find_elements(By.TAG_NAME, "td")
            holders.append({
                "Имотен лист": hc[0].text,
                "Име и презиме": hc[1].text,
                "Град": hc[2].text,
                "Улица": hc[3].text,
                "Број": hc[4].text,
                "Дел на посед": hc[5].text,
            })
        data["Носители на право"] = holders

        if len(driver.window_handles) > 1:
            driver.close()
            driver.switch_to.window(original_handle)
        else:
            driver.back()

        result.append(data)
        wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tr.parcels-table-body-row")))

    return result


# -----------------------------
# Excel writer (single sheet, hierarchical)
# -----------------------------
def write_results_to_excel(region_results, filename="results.xlsx"):
    """
    Single-sheet, hierarchical view with indentation and wrapping.
    Region rows are bold and shaded; parcel rows indented; holder rows more indented.
    Includes input_katastar and not-found messages.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    # Header row
    ws.append([
        "Input Region", "Input KatastarRegion (optional)", "Input Parcel",
        "Kat. Odd. (Region Suggestion)",
        "Имотен лист", "Број/дел", "Култура", "Површина m2", "Место", "Право",
        "Име и презиме", "Град", "Улица", "Број", "Дел на посед",
        "Note"
    ])
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    # Styles
    region_font = Font(bold=True)
    region_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    region_alignment = Alignment(wrap_text=True)
    parcel_alignment = Alignment(indent=1, wrap_text=True)
    holder_alignment = Alignment(indent=2, wrap_text=True)

    row_num = 2
    for entry in region_results:
        region_name = entry["region_name"]
        input_region = entry.get("input_region")
        input_katastar = entry.get("input_katastar")  # NEW
        input_parcel = entry.get("input_parcel")
        note = entry.get("note", "")

        # Region row
        ws.append([
            input_region, input_katastar, input_parcel,
            region_name,
            "", "", "", "", "", "",
            "", "", "", "", "",
            note
        ])
        for cell in ws[row_num]:
            cell.font = region_font
            cell.fill = region_fill
            cell.alignment = region_alignment
        row_num += 1

        # If we had a "not found" case for this region suggestion, no parcels
        if note:
            continue

        # Parcel rows + holders
        for parcel in entry["parcels"]:
            ws.append([
                "", "", "",      # keep region context empty, use indent
                "",              # region suggestion column empty for parcel rows
                parcel["Имотен лист"],
                parcel["Број/дел"],
                parcel["Култура"],
                parcel["Површина m2"],
                parcel["Место"],
                parcel["Право"],
                "", "", "", "", "",  # holder columns empty for parcel line
                ""
            ])
            # Indent the first 10 columns of the parcel row
            for cell in ws[row_num][0:10]:
                cell.alignment = parcel_alignment
            row_num += 1

            for holder in parcel["Носители на право"]:
                ws.append([
                    "", "", "", "",  # keep context empty
                    "", "", "", "", "", "",  # parcel columns empty for holder line
                    holder["Име и презиме"],
                    holder["Град"],
                    holder["Улица"],
                    holder["Број"],
                    holder["Дел на посед"],
                    ""
                ])
                for cell in ws[row_num]:
                    cell.alignment = holder_alignment
                row_num += 1

    # Autosize columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                text = str(cell.value)
                # For wrapped cells, a reasonable width; still base on longest line
                longest_line = max(len(line) for line in text.splitlines())
                max_length = max(max_length, longest_line)
        ws.column_dimensions[column].width = min(max_length + 2, 60)  # cap width to keep sheet readable

    wb.save(filename)


# -----------------------------
# Core scrape flow (with optional katastar region)
# -----------------------------
def scrape_katastar(region: str, parcel: str, katastar_region: Optional[str] = None):
    """
    - If katastar_region is provided:
        * Type only 'region' to get suggestions.
        * If any suggestion's text contains katastar_region (case-insensitive substring), click that ONE and scrape.
        * If none match, return a single entry with a 'note' explaining not found.
    - If katastar_region is NOT provided:
        * Iterate all suggestions (existing behavior) and scrape each.
    """
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--window-size=1920,1080")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    wait = WebDriverWait(driver, 20)

    region_results = []
    try:
        driver.get("https://e-uslugi.katastar.gov.mk/")

        suggestions = get_region_suggestions(driver, wait, region)
        if not suggestions:
            # No suggestions at all for base region
            region_results.append({
                "region_name": "(no suggestions)",
                "parcels": [],
                "input_region": region,
                "input_katastar": katastar_region,
                "input_parcel": parcel,
                "note": f"No region suggestions found for '{region}'."
            })
            return region_results

        # Normalize texts for matching
        def norm(s: str) -> str:
            return s.strip().lower()

        if katastar_region:
            target = norm(katastar_region)
            # Refresh suggestions after typing region
            suggestions = get_region_suggestions(driver, wait, region)
            # Find a suggestion that contains the katastar part (e.g. '... - СКОПЈЕ')
            match_el = None
            match_text = None
            for el in suggestions:
                text = el.text or ""
                if target in norm(text):
                    match_el = el
                    match_text = text
                    break

            if not match_el:
                # None matched — record a not-found entry
                region_results.append({
                    "region_name": f"(no match among suggestions for '{region}')",
                    "parcels": [],
                    "input_region": region,
                    "input_katastar": katastar_region,
                    "input_parcel": parcel,
                    "note": f"Katastar region '{katastar_region}' not found in suggestions."
                })
                return region_results

            # Click only the matched suggestion, then proceed as usual
            region_name = match_text
            driver.execute_script("arguments[0].click();", match_el)
            region_entry = {
                "region_name": region_name,
                "parcels": [],
                "input_region": region,
                "input_katastar": katastar_region,
                "input_parcel": parcel
            }

            if select_parcel(driver, wait, parcel):
                region_entry["parcels"] = extract_parcel_and_holders(driver, wait)
            else:
                region_entry["note"] = f"No parcel suggestions found for '{parcel}' in region '{region_name}'."

            region_results.append(region_entry)
            return region_results

        else:
            # No katastar filter: iterate through ALL suggestions (existing behavior)
            for idx in range(len(suggestions)):
                suggestions = get_region_suggestions(driver, wait, region)
                current = suggestions[idx]
                region_name = current.text
                driver.execute_script("arguments[0].click();", current)

                region_entry = {
                    "region_name": region_name,
                    "parcels": [],
                    "input_region": region,
                    "input_katastar": None,
                    "input_parcel": parcel
                }

                if select_parcel(driver, wait, parcel):
                    region_entry["parcels"] = extract_parcel_and_holders(driver, wait)
                else:
                    region_entry["note"] = f"No parcel suggestions found for '{parcel}' in region '{region_name}'."

                region_results.append(region_entry)
                driver.get("https://e-uslugi.katastar.gov.mk/")

            return region_results

    finally:
        driver.quit()


# -----------------------------
# Input parsing (2 or 3 items per line)
# -----------------------------
def read_input_file(file_path: str) -> List[Tuple[str, Optional[str], str]]:
    """
    Accepts lines in one of two formats (comma/semicolon/tab separated):
      1) Region, Parcel
      2) Region, KatastarRegion, Parcel

    Ignores blank lines and lines starting with '#'.
    Returns list of tuples: (region, katastar_region_or_None, parcel)
    """
    triples: List[Tuple[str, Optional[str], str]] = []
    with open(file_path, "r", encoding="utf-8") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            parts = [p.strip() for p in re.split(r"[,\t;]", line) if p.strip()]
            if len(parts) == 2:
                region, parcel = parts
                triples.append((region, None, parcel))
            elif len(parts) >= 3:
                region, katastar, parcel = parts[0], parts[1], parts[2]
                triples.append((region, katastar, parcel))
            # else: ignore malformed lines silently (or print warning if you prefer)
    return triples


# -----------------------------
# CLI
# -----------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Scrape parcel/right-holder data from e-uslugi.katastar.gov.mk")
    parser.add_argument("--input-file", "-i", help="Path to a file containing Region[,KatastarRegion],Parcel lines")
    parser.add_argument("region", nargs="?", help="Base region text to search (used if no --input-file)")
    parser.add_argument("parcel", nargs="?", help="Parcel to search (used if no --input-file)")
    parser.add_argument("--katastar", "-k", nargs="?", default=None,
                        help="Optional Katastar Region (e.g., 'Скопје'). If provided, only that dropdown option is scraped.")
    args = parser.parse_args()

    jobs: List[Tuple[str, Optional[str], str]] = []
    if args.input_file:
        jobs = read_input_file(args.input_file)
        if not jobs:
            print(f"No valid entries found in input file {args.input_file}")
            sys.exit(1)
    else:
        if not args.region or not args.parcel:
            parser.error("Either --input-file or both positional arguments <region> <parcel> are required.")
        jobs = [(args.region, args.katastar, args.parcel)]

    all_results = []
    for region, katastar_region, parcel in jobs:
        batch = scrape_katastar(region, parcel, katastar_region=katastar_region)
        all_results.extend(batch)

    if all_results:
        write_results_to_excel(all_results, "results.xlsx")
        print(f"Results written to results.xlsx for {len(all_results)} region suggestion(s) across {len(jobs)} input item(s).")
    else:
        print("No results found.")
# -----------------------------